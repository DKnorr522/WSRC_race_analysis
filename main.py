import streamlit as st
import pandas as pd
from matplotlib import pyplot as plt
import seaborn as sns
from plotly import express as px, graph_objects as go, figure_factory as ff, colors
from plotly.subplots import make_subplots
from math import ceil, floor
import os
import openpyxl


def fetch_excel_file(file_name_func: str):
    try:
        wb_func = openpyxl.load_workbook(file_name_func)
    except FileNotFoundError as err:
        print(f"File not found: {err}")
        wb_func = None
    except InvalidFileException as err:
        print(f"File must be .xlsx: {err}")
        wb_func = None
    
    return wb_func

def load_dataframe(wb_func, event_name_func: str) -> pd.DataFrame:
    try:
        header_row: int = 29
        ws = wb_func[event_name_func]
        headers = [cell.value for cell in ws[29]]
        values = [
            [cell.value for cell in ws[row]]
            for row in range(header_row + 2, ws.max_row + 1)
        ]
        df_func = pd.DataFrame(data=values, columns=headers)
    except KeyError as err:
        print(f"Invalid worksheet name: {err}")
        df_func = None
    return df_func

def clean_dataframe(df_func: pd.DataFrame) -> pd.DataFrame:
    df_func.columns = [
        col.lower().replace(
            " ", "_"
        ).replace(
            "(", ""
        ).replace(
            ")", ""
        ).replace(
            ".", ""
        ).replace(
            "/", "_per_"
        )
        for col in df_func.columns
    ]

    # numeric_cols = [
    #     'interval',
    #     'distance_gps',
    #     'distance_imp',
    #     'speed_gps',
    #     'speed_imp',
    #     'stroke_rate',
    #     'total_strokes',
    #     'distance_per_stroke_gps',
    #     'distance_per_stroke_imp',
    #     'heart_rate',
    #     # 'power',
    #     # 'catch',
    #     # 'slip',
    #     # 'finish',
    #     # 'wash',
    #     # 'force_avg',
    #     # 'work',
    #     # 'force_max',
    #     # 'max_force_angle',
    #     'gps_lat',
    #     'gps_lon'
    # ]
    # df_func[numeric_cols] = df[numeric_cols].apply(pd.to_numeric)
    for col in df_func.columns:
        try:
            df_func[col] = df_func[col].apply(pd.to_numeric)
        except TypeError as err:
            print(f"\t{col} could not convert type: {err}")
        except ValueError as err:
            print(f"{col} has no data: {err}")

    df_func["elapsed_time_sec"] = df_func.elapsed_time.apply(
        lambda x: x.hour*3600 + x.minute*60 + x.second + x.microsecond*1e-6
    )
    df_func["split_sec"] = df_func.split_gps.apply(
        lambda x: x.hour*3600 + x.minute*60 + x.second + x.microsecond*1e-6
    )

    return df_func

def speed_limits(df_func: pd.DataFrame) ->  dict[str, float]:
    limits = {
        'max': 5 * floor((500 / df_func.loc[df_func.total_strokes > 5, :].speed_gps.max())/5),
        'min': 5 * ceil((500 / df_func.loc[df_func.total_strokes > 5, :].speed_gps.min())/5)
    }
    return limits

def add_split_lines(df_func: pd.DataFrame, fig_func):
    limits = speed_limits(df_func)
    current_speed = limits['max']
    while current_speed <= limits['min']:
        speed_str = f"{current_speed//60}:{current_speed - 60*(current_speed//60):02d}"
        txt_loc = "bottom left" if current_speed == limits['max'] else "top left"
        fig_func.add_hline(
            y=500 / current_speed,
            line_dash="dash",
            line_width=1,
            annotation_text=speed_str,
            annotation_position=txt_loc,
            line_color="red"
        )
        current_speed += 5
    fig_func.update_yaxes(showgrid=False)
    return fig_func

def add_quarterly_breakdown(
        df_func: pd.DataFrame,
        fig_func,
        num_start_strokes: int = 5,
        num_high_strokes: int = 5,
        show_fastest_slowest: bool = True):
    high_strokes_first, *_, high_strokes_last = df_func.loc[
        (df_func.total_strokes > num_start_strokes) & (df_func.total_strokes < num_start_strokes + num_high_strokes), :
    ].distance_gps.values
    race_distance = df_func.distance_gps.max()
    quarter_distances = [
        (q / 4) * race_distance
        for q in range(5)
    ]

    fig_func.add_vrect(
        x0=high_strokes_first,
        x1=high_strokes_last,
        fillcolor="blue",
        opacity=0.1
    )

    fig_func.add_vrect(
        x0=quarter_distances[1],
        x1=quarter_distances[2],
        fillcolor="green",
        opacity=0.1
    )

    fig_func.add_vrect(
        x0=quarter_distances[2],
        x1=quarter_distances[3],
        fillcolor="yellow",
        opacity=0.1
    )

    fig_func.add_vrect(
        x0=quarter_distances[3],
        x1=quarter_distances[4],
        fillcolor="maroon",
        opacity=0.1
    )

    if show_fastest_slowest:
        df_temp = df_func.loc[df_func.total_strokes > num_start_strokes, :]

        speed_max_dist = df_temp.loc[df_temp.speed_gps == df_func.speed_gps.max(), :].distance_gps.values
        speed_max = [df_temp.speed_gps.max()] * len(speed_max_dist)
        fig_func.add_trace(go.Scatter(
            x=speed_max_dist,
            y=speed_max,
            mode="markers+text",
            name="Fastest",
            text="Fastest",
            textposition="top center"
        ))

        speed_min_dist = df_temp.loc[df_temp.speed_gps == df_temp.speed_gps.min(), :].distance_gps.values
        speed_min = [df_temp.speed_gps.min()] * len(speed_min_dist)
        fig_func.add_trace(go.Scatter(
            x=speed_min_dist,
            y=speed_min,
            mode="markers+text",
            name="Slowest",
            text="Slowest",
            textposition="bottom center"
        ))

    return fig_func

def create_line_plot_speed_stroke_rate(
        df_func: pd.DataFrame,
        strokes_to_ignore_func: int = 5,
        split_lines_func: bool = False,
        breakdown_func: bool = False):
    """
    Adapted secondary y-axis from:
    https://stackoverflow.com/questions/62853539/how-to-plot-on-secondary-y-axis-with-plotly-express
    User: derflo
    On: 6/14/25
    """

    if strokes_to_ignore_func > 0:
        df_func = df_func.loc[df_func.total_strokes > strokes_to_ignore_func, :]

    all_figs = make_subplots(specs=[[{"secondary_y": True}]])

    fig1_func = px.line(
        df_func,
        x="distance_gps",
        y="speed_gps",
        hover_data=["split_gps", "elapsed_time", "stroke_rate", "distance_per_stroke_gps", "total_strokes"],
        labels=labels_dict,
    )
    fig1_func.update_traces({'name': "Speed"})

    fig2_func = px.line(
        df_func,
        x="distance_gps",
        y="stroke_rate",
        hover_data=["elapsed_time", "stroke_rate", "distance_per_stroke_gps", "total_strokes"],
        labels=labels_dict,
    )
    fig2_func.update_traces({'name': "Stroke Rate"}, yaxis="y2")

    all_figs.add_traces(fig1_func.data + fig2_func.data)
    all_figs.layout.xaxis.title = "Distance (m)"
    all_figs.layout.yaxis.title = "Speed (m/s)"
    all_figs.layout.yaxis2.title = "Stroke Rate"

    all_figs.for_each_trace(lambda t: t.update(
        line=dict(color=t.marker.color),
        showlegend=True
    ))

    if split_lines_func:
        all_figs = add_split_lines(df.copy(), all_figs)
    if breakdown_func:
        all_figs = add_quarterly_breakdown(df.copy(), all_figs)

    all_figs.update_xaxes(range=[0, df_func.distance_gps.max()*1.05])
    return all_figs

def create_scatter_plot_speed_colored_stroke_rate(
        df_func: pd.DataFrame,
        strokes_to_ignore_func: int = 5,
        split_lines_func: bool = False,
        breakdown_func: bool = False):
    if strokes_to_ignore_func > 0:
        df_func = df_func.loc[df_func.total_strokes > strokes_to_ignore_func, :]
    fig = px.scatter(
        df_func,
        x="distance_gps",
        y="speed_gps",
        color="stroke_rate",
        # color="distance_per_stroke_gps",
        hover_data=[
            "split_gps",
            "elapsed_time",
            "stroke_rate",
            "distance_per_stroke_gps",
            "total_strokes"
        ],
        labels=labels_dict,
        color_continuous_scale='aggrnyl',
    )

    if split_lines_func:
        fig = add_split_lines(df.copy(), fig)
    if breakdown_func:
        fig = add_quarterly_breakdown(df.copy(), fig, show_fastest_slowest=False)

    fig.update_xaxes(range=[0, df_func.distance_gps.max()*1.05])

    return fig

def create_scatter_plot_stroke_rate_colored_speed(
        df_func: pd.DataFrame,
        strokes_to_ignore_func: int = 5,
        breakdown_func: bool = False):
    if strokes_to_ignore_func > 0:
        df_func = df_func.loc[df_func.total_strokes > strokes_to_ignore_func, :]
    fig = px.scatter(
        df_func,
        x="distance_gps",
        y="stroke_rate",
        color="speed_gps",
        hover_data=[
            "elapsed_time",
            "stroke_rate",
            "distance_per_stroke_gps",
            "total_strokes",
            "split_gps"
        ],
        labels=labels_dict,
        color_continuous_scale='aggrnyl',
    )

    if breakdown_func:
        fig = add_quarterly_breakdown(df_func.copy(), fig, show_fastest_slowest=False)

    fig.update_xaxes(range=[0, df_func.distance_gps.max()*1.05])
    return fig

def create_box_plot_stroke_rate_speed(
        df_func: pd.DataFrame,
        strokes_to_ignore_func: int = 5,
        split_lines_func: bool = False):
    if strokes_to_ignore_func > 0:
        df_func = df_func.loc[df_func.total_strokes > strokes_to_ignore_func, :]
    fig = px.box(
        df_func,
        x="stroke_rate",
        y="speed_gps",
        labels=labels_dict,
        hover_data=[
            "split_gps"
        ]
    )

    if split_lines_func:
        fig = add_split_lines(df_func.copy(), fig)

    return fig

def plot_course_map(df_func: pd.DataFrame, size: float = 0.1, zoom: int = 14) -> None:
    st.map(
        df_func,
        latitude="gps_lat",
        longitude="gps_lon",
        size=size,
        zoom=zoom
    )
    return


file_name = "2025 Biernacki.xlsx"

st.markdown(
    "<h1 style='text-align: center;'> WSRC Race Results </h1>",
    unsafe_allow_html=True
)
st.divider()

labels_dict = {
    "speed_gps": "Speed (m/s)",
    "split_gps": "Split",
    "distance_gps": "Distance (m)",
    "stroke_rate": "Stroke Rate",
    "elapsed_time": "Time",
    "distance_per_stroke_gps": "Meters per Stroke",
    "total_strokes": "Stroke Count"
}

col_race, col_breakdown = st.columns(2)

with col_race:
    wb = fetch_excel_file(file_name)
    race_choices = wb.sheetnames
    race_choice = st.selectbox(
        "",
        options=race_choices,
        index=None,
        placeholder="Choose a race"
    )

if race_choice:
    with col_breakdown:
        split_lines = st.checkbox(
            "Show splits",
            value=False
        )
        breakdown = st.checkbox(
            "Show quarterly breakdown",
            value=False
        )
        show_start = st.checkbox(
            "Show starting strokes",
            value=False
        )
        strokes_to_ignore = 0 if show_start else 5

    df = load_dataframe(wb, race_choice)
    st.write(df.dtypes)
    st.write(df.columns)
    st.write(df.head())
    st.write(df.drop(columns="").head())
    df = clean_dataframe(df)

    fig1 = create_line_plot_speed_stroke_rate(df.copy(), strokes_to_ignore, split_lines, breakdown)
    st.plotly_chart(fig1)

    fig2 = create_scatter_plot_speed_colored_stroke_rate(df.copy(), strokes_to_ignore, split_lines, breakdown)
    st.plotly_chart(fig2)

    fig3 = create_scatter_plot_stroke_rate_colored_speed(df.copy(), strokes_to_ignore, breakdown)
    st.plotly_chart(fig3)

    fig4 = create_box_plot_stroke_rate_speed(df.copy(), strokes_to_ignore, split_lines)
    st.plotly_chart(fig4)

    plot_course_map(df.copy(), 0.1, 14)

